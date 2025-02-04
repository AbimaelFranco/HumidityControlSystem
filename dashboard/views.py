from django.shortcuts import render
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from django.utils.timezone import now, timedelta, localtime
from django.core.paginator import Paginator
from dashboard.models import Records, configuration
#############################
import subprocess
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import make_aware
import datetime
import pytz



def dashboard_home(request):
    # Obtener registros de las últimas 24 horas
    last_24_hours = now() - timedelta(hours=1)
    records = Records.objects.filter(timestamp__gte=last_24_hours)
    # records = records[::-1]

    try:
        # Organizar datos para la gráfica y convertir de Decimal a float
        timestamps = [localtime(record.timestamp).strftime("%H:%M") for record in records]  # Convertir a zona horaria local
        
        # Convertir a float para evitar problemas con los tipos Decimal
        humidity1 = [float(record.humidity1) for record in records]
        temperature1 = [float(record.temperature1) for record in records]
        humidity2 = [float(record.humidity2) for record in records]
        temperature2 = [float(record.temperature2) for record in records]

        # Valores promedio
        avg_humidity = [float(record.humidity_average) for record in records]
        avg_temperature = [float(record.temperature_average) for record in records]

        # Obtener el último registro
        last_record = records.last()

        # Obtener los últimos valores de temperatura, humedad y la hora
        last_timestamp = localtime(last_record.timestamp).strftime("%H:%M")
        last_avg_humidity = float(last_record.humidity_average)
        last_avg_temperature = float(last_record.temperature_average)

        #Obtener configuraciones
        configurations = configuration.objects.all()
        last_configuration = configurations[0]
        temperature_min_alert = float(last_configuration.temperature_min_alert)
        temperature_max_alert = float(last_configuration.temperature_max_alert)
        humidity_min_alert = float(last_configuration.humidity_min_alert)
        humidity_max_alert = float(last_configuration.humidity_max_alert)

        context = {
        'timestamps': timestamps,
        'humidity1': humidity1,
        'temperature1': temperature1,
        'humidity2': humidity2,
        'temperature2': temperature2,
        'avg_humidity': avg_humidity,
        'avg_temperature': avg_temperature,
        'last_timestamp': last_timestamp,
        'last_avg_humidity': last_avg_humidity,
        'last_avg_temperature': last_avg_temperature,
        'temperature_min_alert':temperature_min_alert,
        'temperature_max_alert':temperature_max_alert,
        'humidity_min_alert':humidity_min_alert,
        'humidity_max_alert':humidity_max_alert,
    }
    
    except:

        context = {
        'timestamps': '0',
        'humidity1': 0,
        'temperature1': 0,
        'humidity2': 0,
        'temperature2': 0,
        'avg_humidity': 0,
        'avg_temperature': 0,
        'last_timestamp': 0,
        'last_avg_humidity': 0,
        'last_avg_temperature': 0,
        'temperature_min_alert':0,
        'temperature_max_alert':0,
        'humidity_min_alert':0,
        'humidity_max_alert':0,
    }  

    # Enviar los datos al template
    return render(request, 'dashboard/home.html', context)

def graphs(request):
    # Obtener registros de las últimas 24 horas
    last_24_hours = now() - timedelta(hours=24)
    records = Records.objects.filter(timestamp__gte=last_24_hours)
    # records = records[::-1]

    try:
        # Organizar datos para la gráfica y convertir de Decimal a float
        timestamps = [localtime(record.timestamp).strftime("%H:%M") for record in records]  # Convertir a zona horaria local
        
        # Convertir a float para evitar problemas con los tipos Decimal
        humidity1 = [float(record.humidity1) for record in records]
        temperature1 = [float(record.temperature1) for record in records]
        humidity2 = [float(record.humidity2) for record in records]
        temperature2 = [float(record.temperature2) for record in records]

        # Valores promedio
        avg_humidity = [float(record.humidity_average) for record in records]
        avg_temperature = [float(record.temperature_average) for record in records]

        # Obtener el último registro
        last_record = records.last()

        # Obtener los últimos valores de temperatura, humedad y la hora
        last_timestamp = localtime(last_record.timestamp).strftime("%H:%M")
        last_avg_humidity = float(last_record.humidity_average)
        last_avg_temperature = float(last_record.temperature_average)


        context = {
            'timestamps': timestamps,
            'humidity1': humidity1,
            'temperature1': temperature1,
            'humidity2': humidity2,
            'temperature2': temperature2,
            'avg_humidity': avg_humidity,
            'avg_temperature': avg_temperature,
            'last_timestamp': last_timestamp,
            'last_avg_humidity': last_avg_humidity,
            'last_avg_temperature': last_avg_temperature,
        }
    except:

        context = {
        'timestamps': '0',
        'humidity1': 0,
        'temperature1': 0,
        'humidity2': 0,
        'temperature2': 0,
        'avg_humidity': 0,
        'avg_temperature': 0,
        'last_timestamp': 0,
        'last_avg_humidity': 0,
        'last_avg_temperature': 0,
        'temperature_min_alert':0,
        'temperature_max_alert':0,
        'humidity_min_alert':0,
        'humidity_max_alert':0,
    }  
    
    # Enviar los datos al template
    return render(request, 'dashboard/graphs.html', context)

def registers(request):
    registros = Records.objects.all().order_by('-timestamp')  # Ordenar por fecha y hora descendente
    paginator = Paginator(registros, 20)  # 20 registros por página

    page_number = request.GET.get('page')  # Obtener el número de página desde los parámetros de la URL
    page_obj = paginator.get_page(page_number)  # Obtener la página actual

    context = {
        'page_obj': page_obj,
    }
    return render(request, 'dashboard/registers.html', context)

def export_to_excel(request):
    # Crear un libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    # Estilo de encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Escribir los encabezados en el archivo Excel (nuevo orden de columnas)
    headers = [
        'ID', 'Timestamp', 'Temperature Sensor 1', 'Temperature Sensor 2', 
        'Temperature Average', 'Humidity Sensor 1', 'Humidity Sensor 2', 
        'Humidity Average'
    ]
    ws.append(headers)

    # Aplicar estilo a los encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Obtener todos los registros
    records = Records.objects.all()

    # Agregar los datos de los registros al archivo Excel (nuevo orden)
    for record in records:
        ws.append([
            record.id, record.timestamp.strftime('%Y-%m-%d %H:%M:%S'), 
            record.temperature1, record.temperature2, record.temperature_average,
            record.humidity1, record.humidity2, record.humidity_average
        ])

    # Configurar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Registros_Humedad_y_Temperatura.xlsx"'

    # Guardar el archivo Excel en la respuesta
    wb.save(response)

    return response


@csrf_exempt  # Deshabilitar CSRF para pruebas, asegúrate de manejar la seguridad adecuadamente
def update_system_time(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            new_time = data.get("datetime")

            if not new_time:
                return JsonResponse({"error": "No datetime provided"}, status=400)

            # Convertir la hora enviada a UTC (si es necesario)
            utc_time = datetime.datetime.strptime(new_time, "%Y-%m-%d %H:%M:%S")

            # Convertir la hora a la zona horaria de Guatemala (CST)
            guatemala_tz = pytz.timezone('America/Guatemala')
            aware_time = pytz.utc.localize(utc_time)  # Convertir a hora UTC (aware time)
            guatemala_time = aware_time.astimezone(guatemala_tz)  # Convertir a la zona horaria de Guatemala

            # Formatear la hora para configurarla en la Raspberry Pi
            formatted_time = guatemala_time.strftime("%Y-%m-%d %H:%M:%S")

            # Comando para actualizar la hora en la Raspberry Pi
            subprocess.run(["sudo", "timedatectl", "set-time", formatted_time], check=True)

            return JsonResponse({"success": True, "message": f"Time updated to {guatemala_time.strftime('%Y-%m-%d %H:%M:%S')}"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
